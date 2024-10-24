import sys
from datetime import datetime
import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
from tqdm import tqdm

from modules.vpc import load_and_transform_vpc_data
from modules.tgw import load_and_transform_tgw_data
from modules.vep import load_and_transform_vep_data
from modules.pc import load_and_transform_pc_data
from modules.subnet import load_and_transform_subnet_data
from modules.security_groups import load_and_transform_security_groups_data
from modules.autoscaling import load_and_transform_autoscaling_data
from modules.cloudfront import load_and_transform_cloudfront_data
from modules.docdb import load_and_transform_docdb_data
from modules.ec import load_and_transform_elasticache_data
from modules.ec2 import load_and_transform_ec2_data
from modules.elb import load_and_transform_elb_data
from modules.iamrole import load_and_transform_iam_role_data
from modules.iamuser import load_and_transform_iam_user_data
from modules.iamgroup import load_and_transform_iam_group_data
from modules.nacl import load_and_transform_nacl_data
from modules.s3 import load_and_transform_s3_data
from modules.tg import load_and_transform_target_group_data
from modules.rds import load_and_transform_rds_data

# Prompt the user for PostgreSQL connection details
host = '127.0.0.1'
port = '9193'
username = 'steampipe'
password = os.getenv('st_password')
db_name = 'steampipe'

def extract_connections(file_path):
    connections = []
    try:
        # Open the file
        with open(file_path, "r") as file:
            content = file.read()
            # Use regular expressions to extract connection names that are not of type "aggregator"
            matches = re.findall(r'connection\s+"([^"]+)"\s*{(?:(?!type\s*=\s*"aggregator").)*}', content, re.DOTALL)
            if matches:
                connections.extend(matches)
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    return connections

schemas = extract_connections(os.path.expanduser("~/.steampipe/config/aws.spc"))
print(schemas)

# Construct the database URI
DB_URI = f'postgresql://{username}:{password}@{host}:{port}/{db_name}'

try:
    engine = create_engine(DB_URI)
    connection = engine.connect()
    print("Connection successful.")

except Exception as e:
    print("Please check your SSO or IAM permissions.")
    sys.exit(1)

# Set the download path
today = datetime.today().strftime('%y%m%d')


def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        column_name = column_cells[0].value

        for cell in column_cells:
            try:
                if cell.value:
                    cell_value = str(cell.value).split("\n")
                    max_cell_length = max(len(line) for line in cell_value)
                    max_length = max(max_length, max_cell_length)
            except Exception as e:
                print(f"Error calculating length for cell {cell.coordinate}: {e}")

        if column_name == 'Tag':
            adjusted_width = 50
        elif column_name == 'Listener From':
            adjusted_width = 15
        elif column_name == 'Listener To':
            adjusted_width = 30
        else:
            adjusted_width = max_length + 2

        sheet.column_dimensions[column_letter].width = adjusted_width


def style_header(sheet):
    header_fill = PatternFill(start_color="334d1d", end_color="334d1d", fill_type="solid")
    header_font = Font(color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center')


def apply_borders(sheet):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def center_align_cells(sheet):
    for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=1), start=1):
        if col[0].value == "Tag":
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(vertical='center')


def process_and_save_sheets(queries, output_excel_path):
    try:
        data_dict = {}

        for query in tqdm(queries, desc="Executing SQL Queries"):
            data_dict[f"{query}_data"] = pd.read_sql(queries[query], engine)

        with pd.ExcelWriter(output_excel_path, engine='xlsxwriter') as writer:
            transformation_tasks = [
                ('VPC', load_and_transform_vpc_data,
                 [data_dict['vpc_data'], data_dict['igw_data'], data_dict['ngw_data']]),
                ('VPC Endpoint', load_and_transform_vep_data, [data_dict['vep_data']]),
                ('Peering Connection', load_and_transform_pc_data, [data_dict['pc_data']]),
                ('Transit Gateway', load_and_transform_tgw_data, [data_dict['tgw_data']]),
                ('Subnet', load_and_transform_subnet_data,
                 [data_dict['subnet_data'], data_dict['rt_data'], data_dict['nacl_data']]),
                ('Security Groups', load_and_transform_security_groups_data,
                 [data_dict['sg_data'], data_dict['sgrule_data']]),
                ('Network ACLs', load_and_transform_nacl_data, [data_dict['nacl_data']]),
                ('EC2', load_and_transform_ec2_data, [data_dict['ec2_data'], data_dict['ebs_data']]),
                ('ELB', load_and_transform_elb_data,
                 [data_dict['alb_data'], data_dict['nlb_data'], data_dict['lbl_data']]),
                ('Target Group', load_and_transform_target_group_data,
                 [data_dict['tg_data'], data_dict['autoscaling_data'], data_dict['ec2_data']]),
                ('Auto Scaling', load_and_transform_autoscaling_data, [data_dict['autoscaling_data']]),
                ('ElastiCache', load_and_transform_elasticache_data, [data_dict['ec_data'], data_dict['ecrep_data']]),
                ('CloudFront', load_and_transform_cloudfront_data, [data_dict['cloudfront_data']]),
                ('S3', load_and_transform_s3_data, [data_dict['s3_data']]),
                ('IAM Group', load_and_transform_iam_group_data, [data_dict['iamgroup_data']]),
                ('IAM Role', load_and_transform_iam_role_data, [data_dict['iamrole_data']]),
                ('IAM User', load_and_transform_iam_user_data, [data_dict['iamuser_data']]),
                ('RDS', load_and_transform_rds_data,
                 [data_dict['rdscluster_data'], data_dict['rdsinstance_data'], data_dict['cloudwatch_data']]),
                ('DocumentDB', load_and_transform_docdb_data,
                 [data_dict['docdbcluster_data'], data_dict['docdbinstance_data'], data_dict['cloudwatch_data']])
            ]

            for sheet_name, transform_function, params in tqdm(transformation_tasks,
                                                               desc="Transforming and Saving Data"):
                if not params[0].empty:
                    transformed_data = transform_function(*params)
                    transformed_data.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = openpyxl.load_workbook(output_excel_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            adjust_column_widths(ws)
            style_header(ws)
            center_align_cells(ws)
            apply_borders(ws)
        wb.save(output_excel_path)
        print(f"Excel file saved at {output_excel_path}")

    except Exception as e:
        print(f"__init__.py > process_and_save_sheets(): {e}")

# Generate Excel files for each schema
if __name__ == '__main__':
    for schema in schemas:
        output_excel_path = os.path.join('inventory', f'{schema}_inventory_{today}.xlsx')
        os.makedirs('inventory', exist_ok=True)

        queries = {
            'alb': f'SELECT * FROM {schema}.aws_ec2_application_load_balancer',
            'autoscaling': f'SELECT * FROM {schema}.aws_ec2_autoscaling_group',
            'cloudfront': f'SELECT * FROM {schema}.aws_cloudfront_distribution',
            'cloudwatch': f'SELECT * FROM {schema}.aws_cloudwatch_metric',
            'docdbcluster': f'SELECT * FROM {schema}.aws_docdb_cluster',
            'docdbinstance': f'SELECT * FROM {schema}.aws_docdb_cluster_instance',
            'ebs': f'SELECT * FROM {schema}.aws_ebs_volume',
            'ec': f'SELECT * FROM {schema}.aws_elasticache_cluster',
            'ecrep': f'SELECT * FROM {schema}.aws_elasticache_replication_group',
            'ec2': f'SELECT * FROM {schema}.aws_ec2_instance',
            'igw': f'SELECT * FROM {schema}.aws_vpc_internet_gateway',
            'iamgroup': f'SELECT * FROM {schema}.aws_iam_group',
            'iamrole': f'SELECT * FROM {schema}.aws_iam_role',
            'iamuser': f'SELECT * FROM {schema}.aws_iam_user',
            'lbl': f'SELECT * FROM {schema}.aws_ec2_load_balancer_listener',
            'ngw': f'SELECT * FROM {schema}.aws_vpc_nat_gateway',
            'nacl': f'SELECT * FROM {schema}.aws_vpc_network_acl',
            'nlb': f'SELECT * FROM {schema}.aws_ec2_network_load_balancer',
            'pc': f'SELECT * FROM {schema}.aws_vpc_peering_connection',
            'rdscluster': f'SELECT * FROM {schema}.aws_rds_db_cluster',
            'rdsinstance': f'SELECT * FROM {schema}.aws_rds_db_instance',
            'rt': f'SELECT * FROM {schema}.aws_vpc_route_table',
            'sg': f'SELECT * FROM {schema}.aws_vpc_security_group',
            'sgrule': f'SELECT * FROM {schema}.aws_vpc_security_group_rule',
            's3': f'SELECT * FROM {schema}.aws_s3_bucket',
            'subnet': f'SELECT * FROM {schema}.aws_vpc_subnet',
            'tg': f'SELECT * FROM {schema}.aws_ec2_target_group',
            'tgw': f'SELECT * FROM {schema}.aws_ec2_transit_gateway',
            'vep': f'SELECT * FROM {schema}.aws_vpc_endpoint',
            'vpc': f'SELECT * FROM {schema}.aws_vpc',
        }

        sg_queries = [
            "SELECT security_groups AS security_groups FROM aws_dax_cluster",
            "SELECT security_groups AS security_groups FROM aws_ec2_network_load_balancer",
            "SELECT security_groups AS security_groups FROM aws_ec2_launch_configuration",
            "SELECT security_groups AS security_groups FROM aws_ec2_instance",
            "SELECT security_groups AS security_groups FROM aws_ec2_classic_load_balancer",
            "SELECT security_groups AS security_groups FROM aws_ec2_application_load_balancer",
            "SELECT security_groups AS security_groups FROM aws_sagemaker_notebook_instance",
            "SELECT security_groups AS security_groups FROM aws_elasticache_cluster",
            "SELECT security_groups AS security_groups FROM aws_mq_broker",
            "SELECT security_groups AS security_groups FROM aws_efs_mount_target",
            "SELECT security_groups AS security_groups FROM aws_ec2_gateway_load_balancer",
            "SELECT vpc_security_groups AS security_groups FROM aws_dms_replication_instance",
            "SELECT vpc_security_groups AS security_groups FROM aws_redshift_cluster",
            "SELECT vpc_security_groups AS security_groups FROM aws_rds_db_instance",
            "SELECT vpc_security_groups AS security_groups FROM aws_docdb_cluster_instance",
            "SELECT vpc_security_groups AS security_groups FROM aws_rds_db_cluster",
            "SELECT vpc_security_groups AS security_groups FROM aws_docdb_cluster",
            "SELECT vpc_security_groups AS security_groups FROM aws_neptune_db_cluster",
            "SELECT vpc_security_group_ids AS security_groups FROM aws_lambda_version",
            "SELECT vpc_security_group_ids AS security_groups FROM aws_rds_db_proxy",
            "SELECT vpc_security_group_ids AS security_groups FROM aws_lambda_function",
            "SELECT resources_vpc_config AS security_groups FROM aws_eks_cluster",
            "SELECT vpc_settings AS security_groups FROM aws_directory_service_directory",
            "SELECT workspace_security_group_id AS security_groups FROM aws_workspaces_directory"
        ]

        process_and_save_sheets(queries, output_excel_path)
